import os
import random
import time
import csv
import requests
from web3 import Web3, HTTPProvider
from openpyxl import Workbook, load_workbook
import json

# 1. Конфигурационные параметры
MIN_TRANSACTION_DELAY = 30
MAX_TRANSACTION_DELAY = 100
MIN_ACCOUNT_DELAY = 600
MAX_ACCOUNT_DELAY = 900
MIN_TARGET_TXNS = 5
MAX_TARGET_TXNS = 15
USE_PROXIES = True
SHUFFLE_ACCOUNTS = True
EXCEL_PATH = "data.xlsx"

CONTRACTS_AND_QUANTITIES = [
    {"address": "0x4de73D198598C3B4942E95657a12cBc399E4aDB5", "quantity": 1},
    {"address": "0x53cb0B849491590CaB2cc44AF8c20e68e21fc36D", "quantity": 3},
    {"address": "0xca5F4088c11B51c5D2B9FE5e5Bc11F1aff2C4dA7", "quantity": 2},
    {"address": "0x266b7E8Df0368Dd4006bE5469DD4EE13EA53d3a4", "quantity": 3},
    {"address": "0xCc4FF6BB314055846e46490B966745E869546B4a", "quantity": 100},
    {"address": "0x9eAE90902a68584E93a83D7638D3a95ac67FC446", "quantity": 3},
    {"address": "0x4073a52A3fc328D489534Ab908347eC1FcB18f7f", "quantity": 3},
    {"address": "0x12B93dA6865B035AE7151067C8d264Af2ae4be8E", "quantity": 10},
    {"address": "0x266b7E8Df0368Dd4006bE5469DD4EE13EA53d3a4", "quantity": 3},
    {"address": "0xC47ADb3e5dC59FC3B41d92205ABa356830b44a93", "quantity": 2},
    {"address": "0xDcFB6cB9512E50dC54160cB98E5a00B3383F6A53", "quantity": 100}
]

# 2. Вспомогательные функции
def create_excel_file(file_path, private_keys):
    wb = Workbook()
    ws = wb.active

    proxies_list = []
    if USE_PROXIES:
        with open('proxies.txt', 'r') as f:
            proxies_list = [line.strip() for line in f.readlines()]
            proxies_list.extend(proxies_list * (len(private_keys) // len(proxies_list)))

    for idx, (pkey, proxy) in enumerate(zip(private_keys, proxies_list), 2):
        ws.cell(row=idx, column=1, value=pkey)
        ws.cell(row=idx, column=2, value=proxy)
        target_txns = random.randint(MIN_TARGET_TXNS, MAX_TARGET_TXNS)
        ws.cell(row=idx, column=3, value=target_txns)

    wb.save(file_path)


def load_excel_data(file_path, private_keys):
    if not os.path.exists(file_path):
        create_excel_file(file_path, private_keys)

    wb = load_workbook(file_path)
    ws = wb.active

    private_keys = [row[0].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1)]
    proxies_list = [row[1].value for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2)]

    return private_keys, proxies_list


# 3. Основная логика
private_keys = []  # Этот список нужно определить или загрузить где-то ранее в коде
private_keys, proxies_list = load_excel_data(EXCEL_PATH, private_keys)

if SHUFFLE_ACCOUNTS:
    combined = list(zip(private_keys, proxies_list))
    random.shuffle(combined)
    private_keys, proxies_list = zip(*combined)

w3 = Web3(HTTPProvider('https://rpc.zora.energy'))

if not w3.is_connected():
    print("Not connected to RPC")
    exit(1)

print("Connected to RPC")

with open('abi.json', 'r') as f:
    abi = json.load(f)


for idx, (private_key, proxy) in enumerate(zip(private_keys, proxies_list), 1):
    session = requests.Session()
    if USE_PROXIES:
        credentials, ip_port = proxy.split('@')
        session.proxies = {
            "http": f"http://{credentials}@{ip_port}",
            "https": f"http://{credentials}@{ip_port}",
        }
    w3.provider = HTTPProvider('https://rpc.zora.energy', session=session)
    sender_address = w3.eth.account.from_key(private_key).address

    try:
        contract_data = random.choice(CONTRACTS_AND_QUANTITIES)
        contract = w3.eth.contract(address=contract_data["address"],
                                   abi=abi)
        transaction_data = contract.functions.mint(contract_data["quantity"])._encode_transaction_data()
        repeat_mints = random.randint(MIN_TARGET_TXNS, MAX_TARGET_TXNS)
        account_delay = random.randint(MIN_ACCOUNT_DELAY, MAX_ACCOUNT_DELAY)

        for _ in range(repeat_mints):
            txn_params = {
                'to': contract_data["address"],
                'from': sender_address,
                'data': transaction_data
            }
            estimated_gas = w3.eth.estimate_gas(txn_params)
            transaction = {
                'to': contract_data["address"],
                'value': 0,
                'gas': estimated_gas,
                'maxPriorityFeePerGas': int(w3.to_wei('0.0005', 'gwei')),
                'maxFeePerGas': int(w3.to_wei('0.0005', 'gwei')),
                'nonce': w3.eth.get_transaction_count(sender_address),
                'data': transaction_data,
                'chainId': 7777777
            }
            signed_txn = w3.eth.account.sign_transaction(transaction, private_key)
            txn_hash = w3.eth.send_raw_transaction(signed_txn.rawTransaction)
            receipt = w3.eth.wait_for_transaction_receipt(txn_hash)

            status = "Successful" if receipt.status == 1 else "Failed"
            with open('transactions.csv', 'a', newline='') as csvfile:
                csvwriter = csv.writer(csvfile)
                csvwriter.writerow([
                    sender_address,
                    contract_data["address"],
                    f"https://explorer.zora.energy/tx/{txn_hash.hex()}",
                    status
                ])

            print(f"Wallet: {sender_address} [{idx}/{len(private_keys)}]")
            if receipt.status == 1:
                print(f"Transaction successfully completed: https://explorer.zora.energy/tx/{txn_hash.hex()}")
            else:
                print(f"Transaction was rejected: https://explorer.zora.energy/tx/{txn_hash.hex()}")

            delay = random.randint(MIN_TRANSACTION_DELAY, MAX_TRANSACTION_DELAY)
            time.sleep(delay)

        time.sleep(account_delay)

    except Exception as e:
        print(f"Error processing wallet {sender_address}: {e}")